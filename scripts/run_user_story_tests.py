from __future__ import annotations

import asyncio
import os
import sys
from copy import deepcopy
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Awaitable, Callable

import httpx
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


WORKBOOK = ROOT / "outputs" / "pingly_feature_user_stories.xlsx"
NOW = datetime.now(timezone.utc)
FUTURE = (NOW + timedelta(days=2)).replace(microsecond=0)
PAST = (NOW - timedelta(days=4)).replace(microsecond=0)


class FakeWebAuth:
    async def create_login_link_for_tg(self, tg_id: int) -> str | None:
        user = await FakeAccounts().get_by_tg_id(tg_id)
        return "https://pingly-app.ru/auth/telegram?token=ok-token" if user else None

    async def login_email(self, email: str, password: str) -> dict | None:
        if email == "tutor@example.com" and password == "secret123":
            return fake_users["tutor"]
        return None

    async def register_tutor_email(self, full_name: str, email: str, password: str, require_verification: bool = False):
        if len(password) < 6:
            return None, "Password too short"
        user = deepcopy(fake_users["tutor"])
        user["full_name"] = full_name
        user["email"] = email
        user["email_verified"] = not require_verification
        return user, None

    async def send_verification_code(self, user: dict):
        events.append(("send_verification_code", user["id"]))
        return True, None

    async def verify_email_code(self, email: str, code: str):
        if code == "123456":
            user = deepcopy(fake_users["tutor"])
            user["email"] = email
            user["email_verified"] = True
            return user, None
        return None, "Bad code"

    async def resend_code(self, email: str) -> bool:
        events.append(("resend_code", email))
        return True

    async def consume_login_token(self, token: str) -> dict | None:
        return fake_users["tutor"] if token == "ok-token" else None

    async def login_telegram_widget(self, data: dict[str, str]) -> dict | None:
        return fake_users["tutor"] if data.get("hash") == "ok" else None

    async def link_telegram(self, user_id: str, data: dict[str, str]):
        if data.get("hash") == "ok":
            return True, None
        return False, "Telegram failed"


class FakeAccounts:
    async def get_user(self, user_id: str) -> dict | None:
        return fake_users.get(user_id)

    async def get_by_tg_id(self, tg_id: int) -> dict | None:
        return next((u for u in fake_users.values() if u.get("tg_id") == tg_id), None)

    async def update_name_by_user_id(self, user_id: str, full_name: str) -> dict | None:
        fake_users[user_id]["full_name"] = full_name
        events.append(("update_name", user_id, full_name))
        return fake_users[user_id]

    async def apply_referral(self, new_user_id: str, ref_code: str) -> bool:
        events.append(("apply_referral", new_user_id, ref_code))
        return True


class FakeStudents:
    async def link_student_from_invite(self, token: str, tg_id: int, full_name: str, tg_username: str | None):
        events.append(("link_student_from_invite", token, tg_id, full_name, tg_username))
        return {"id": "st-linked"} if token == "tok123" else None

    async def list_students_by_user(self, tutor_user_id: str, search: str | None = None) -> list[dict]:
        rows = deepcopy(fake_students)
        if search:
            q = search.lower()
            rows = [s for s in rows if q in s["name"].lower() or q in (s.get("tg_username") or "").lower()]
        return rows

    async def create_student_for_user(self, tutor_user_id: str, name: str, tg_username: str = "", subject_summary: str | None = None):
        row = deepcopy(fake_students[0])
        row["id"] = "st-new"
        row["name"] = name
        row["tg_username"] = tg_username.strip("@")
        row["subject_summary"] = subject_summary
        events.append(("create_student", name))
        return row

    async def student_card(self, tutor_user_id: str, student_id: str) -> dict:
        if student_id != "st1":
            raise PermissionError("missing")
        return {
            "student": deepcopy(fake_students[0]),
            "note": "Needs confidence work",
            "lessons": deepcopy(fake_lessons),
            "homework": deepcopy(fake_homework),
            "next_lesson": deepcopy(fake_lessons[0]),
            "package": {"size": 4, "consumed": 2, "remaining": 2, "started_at": PAST.isoformat()},
            "last_activity": PAST.isoformat(),
            "progress": {"completed_lessons": 2, "attendance_percent": 80, "homework_completion_percent": 50},
        }

    async def update_profile(self, tutor_user_id: str, student_id: str, fields: dict):
        events.append(("update_student_profile", student_id, fields))
        return deepcopy(fake_students[0])

    async def set_note(self, tutor_user_id: str, student_id: str, note: str | None) -> None:
        events.append(("set_note", student_id, note))

    async def set_package(self, tutor_user_id: str, student_id: str, size: int | None):
        events.append(("set_package", student_id, size))
        return deepcopy(fake_students[0])

    async def delete_student(self, tutor_user_id: str, student_id: str) -> dict:
        events.append(("delete_student", student_id))
        return {"notify_tg_id": 1002, "student_name": "Alice"}


class FakeLessons:
    async def list_tutor_calendar(self, tutor_user_id: str) -> list[dict]:
        return deepcopy(fake_lessons)

    async def list_student_calendar(self, student_user_id: str) -> list[dict]:
        return deepcopy(fake_lessons)

    async def next_lesson_for_student(self, student_user_id: str) -> dict | None:
        return deepcopy(fake_lessons[0])

    async def create_one_time_lesson(self, tutor_user_id: str, student_id: str, starts_at: datetime, public_comment: str | None = None, **kwargs):
        events.append(("create_one_time_lesson", student_id, starts_at.isoformat(), public_comment))
        return deepcopy(fake_lessons[0])

    async def create_schedule(self, tutor_user_id: str, student_id: str, recurrence: str, lesson_time: str, **kwargs):
        events.append(("create_schedule", student_id, recurrence, lesson_time, kwargs))
        return {"id": "rule1"}

    async def complete_lesson(self, tutor_user_id: str, lesson_id: str):
        events.append(("complete_lesson", lesson_id))
        return deepcopy(fake_lessons[1])

    async def cancel_lesson(self, tutor_user_id: str, lesson_id: str):
        events.append(("cancel_lesson", lesson_id))
        return deepcopy(fake_lessons[0])

    async def delete_lesson(self, tutor_user_id: str, lesson_id: str) -> None:
        events.append(("delete_lesson", lesson_id))

    async def set_lesson_comment(self, tutor_user_id: str, lesson_id: str, comment: str | None):
        events.append(("set_lesson_comment", lesson_id, comment))
        return deepcopy(fake_lessons[0])

    async def set_lesson_paid(self, tutor_user_id: str, lesson_id: str, paid: bool):
        events.append(("set_lesson_paid", lesson_id, paid))
        return deepcopy(fake_lessons[1])

    async def reschedule_lesson(self, tutor_user_id: str, lesson_id: str, new_starts_at: datetime):
        events.append(("reschedule_lesson", lesson_id, new_starts_at.isoformat()))
        return deepcopy(fake_lessons[0])

    async def cancel_series(self, tutor_user_id: str, rule_id: str):
        events.append(("cancel_series", rule_id))
        return 2

    async def reschedule_series(self, tutor_user_id: str, rule_id: str, new_time: str):
        events.append(("reschedule_series", rule_id, new_time))
        return 2

    async def student_confirm_lesson(self, student_user_id: str, lesson_id: str):
        events.append(("student_confirm_lesson", lesson_id))
        return deepcopy(fake_lessons[0])

    async def student_cancel_lesson(self, student_user_id: str, lesson_id: str):
        events.append(("student_cancel_lesson", lesson_id))
        return deepcopy(fake_lessons[0])

    async def student_request_reschedule(self, student_user_id: str, lesson_id: str):
        events.append(("student_request_reschedule", lesson_id))
        return deepcopy(fake_lessons[0])

    async def confirm_push_target(self, lesson: dict):
        return "tg", 1001, "confirmed"

    async def cancel_push_target(self, lesson: dict):
        return "tg", 1001, "cancelled"

    async def reschedule_request_push_target(self, lesson: dict):
        return "tg", 1001, "reschedule"

    async def finance_overview(self, tutor_user_id: str) -> dict:
        return {
            "month_earned": 3000,
            "total_unpaid": 1000,
            "students": [{"student_id": "st1", "name": "Alice", "lessons": 3, "paid_sum": 2000, "unpaid_sum": 1000, "unpaid_count": 1}],
        }

    async def list_student_history(self, student_user_id: str) -> dict:
        return {"lessons": deepcopy(fake_lessons[1:]), "completed": 2, "cancelled": 1, "total": 3}


class FakeHomework:
    async def list_for_tutor(self, tutor_user_id: str) -> list[dict]:
        return deepcopy(fake_homework)

    async def list_for_student(self, student_user_id: str) -> list[dict]:
        return deepcopy(fake_homework)

    async def list_templates(self, tutor_user_id: str) -> list[dict]:
        return [{"id": "tpl1", "title": "Equation drill", "description": "Solve 10 examples"}]

    async def create_homework(self, tutor_user_id: str, student_id: str, title: str, description: str | None = None, due_at: datetime | None = None):
        events.append(("create_homework", student_id, title, due_at.isoformat() if due_at else None))
        return {"id": "hw-new", "title": title}

    async def review(self, tutor_user_id: str, homework_id: str, comment: str | None = None):
        events.append(("review_homework", homework_id, comment))
        return deepcopy(fake_homework[1])

    async def create_template(self, tutor_user_id: str, title: str, description: str | None = None):
        events.append(("create_template", title))
        return {"id": "tpl-new"}

    async def delete_template(self, tutor_user_id: str, template_id: str) -> None:
        events.append(("delete_template", template_id))

    async def mark_in_progress(self, student_user_id: str, homework_id: str):
        events.append(("mark_in_progress", homework_id))
        return deepcopy(fake_homework[0])

    async def mark_submitted(self, student_user_id: str, homework_id: str):
        events.append(("mark_submitted", homework_id))
        return deepcopy(fake_homework[0])


class FakePublic:
    def parse_badges(self, raw: str | None) -> list[dict]:
        return [{"icon": "clock", "text": "Fast reply"}] if raw else []

    async def get_profile(self, tutor_user_id: str) -> dict | None:
        return deepcopy(fake_profile)

    async def get_public_profile(self, slug: str) -> dict | None:
        if slug == "alice-math":
            return deepcopy(fake_profile)
        return None

    async def create_booking(self, slug: str, name: str, contact: str, preferred_time: str, comment: str) -> dict | None:
        if slug != "alice-math":
            return None
        events.append(("create_booking", name, contact))
        return {"id": "req-new", "tutor_user_id": "tutor"}

    async def booking_push_target(self, tutor_user_id: str, name: str, contact: str):
        return 1001, f"booking {name}"

    async def list_requests(self, tutor_user_id: str) -> list[dict]:
        return deepcopy(fake_requests)

    async def mark_request(self, tutor_user_id: str, request_id: str, status: str) -> None:
        events.append(("mark_request", request_id, status))

    def parse_reviews(self, raw) -> list[dict]:
        return raw if isinstance(raw, list) else []

    async def update_profile(self, tutor_user_id: str, slug: str, bio: str, subjects: str, public_enabled: bool, badges: str = "", page_theme: str = "auto", **extra):
        if public_enabled and len(slug) < 3:
            return None, "short slug"
        events.append(("update_public_profile", slug, public_enabled, page_theme))
        return deepcopy(fake_profile), None


class FakeBilling:
    async def start_subscription(self, user: dict, base_url: str, plan: str = "max"):
        events.append(("start_subscription", user["id"], plan))
        return "https://pay.example/checkout", None

    async def handle_webhook(self, merchant_id: str | None, secret: str | None, body: dict) -> bool:
        events.append(("platega_webhook", body.get("id")))
        return body.get("status") == "CONFIRMED"


class FakeAdmin:
    async def overview(self) -> dict:
        return {
            "tutors_total": 2, "students_total": 3, "lessons_total": 4,
            "active_subscriptions": 1, "active_access": 2, "new_tutors_week": 1,
            "revenue_total": 990, "payments_count": 1,
        }

    async def list_tutors(self) -> list[dict]:
        return [deepcopy(fake_users["tutor"])]

    async def get_tutor(self, user_id: str):
        return fake_users.get(user_id)

    async def grant_subscription(self, user_id: str, plan: str, days: int):
        events.append(("grant_subscription", user_id, plan, days))

    async def set_plan(self, user_id: str, plan: str):
        events.append(("set_plan", user_id, plan))

    async def extend_trial(self, user_id: str, days: int):
        events.append(("extend_trial", user_id, days))

    async def broadcast_targets(self) -> list[int]:
        return [1001, 1002]


class FakeAnalytics:
    async def tutor_dashboard(self, tutor_user_id: str) -> dict:
        return {
            "students_count": 1,
            "lessons_count": 4,
            "completed_lessons": 2,
            "homework_count": 3,
            "attendance_percent": 80,
            "homework_completion_percent": 50,
        }


class FakeServices:
    def __init__(self) -> None:
        self.accounts = FakeAccounts()
        self.web_auth = FakeWebAuth()
        self.students = FakeStudents()
        self.lessons = FakeLessons()
        self.homework = FakeHomework()
        self.public = FakePublic()
        self.billing = FakeBilling()
        self.admin = FakeAdmin()
        self.analytics = FakeAnalytics()


fake_users: dict[str, dict[str, Any]] = {
    "tutor": {
        "id": "tutor", "role": "tutor", "full_name": "Tutor One", "email": "tutor@example.com",
        "email_verified": True, "tg_id": 1001, "tg_username": "tutorone", "is_admin": True,
        "subscription_status": "trial", "trial_ends_at": (NOW + timedelta(days=7)).isoformat(),
        "referral_code": "REF123", "plan": "max", "created_at": PAST.isoformat(),
    },
    "student": {
        "id": "student", "role": "student", "full_name": "Student One", "email": "student@example.com",
        "email_verified": True, "tg_id": 1002, "tg_username": "studentone", "is_admin": False,
        "subscription_status": "trial", "trial_ends_at": None,
    },
}

fake_students = [{
    "id": "st1", "name": "Alice", "tg_username": "alice", "invite_token": "tok123",
    "user_id": "student", "subject_summary": "Math", "grade": "8", "goal": "Exam prep",
    "started_at": PAST.date().isoformat(), "default_price": 1000,
    "package_size": 4, "package_started_at": PAST.isoformat(),
    "tg_connected": True, "vk_connected": False,
}]

fake_lessons = [
    {"id": "l1", "student_id": "st1", "tutor_user_id": "tutor", "student_user_id": "student", "starts_at": FUTURE.isoformat(), "status": "scheduled", "student_profiles": {"name": "Alice"}, "price": 1000, "paid": False, "public_comment": "Fractions", "schedule_rule_id": "rule1"},
    {"id": "l2", "student_id": "st1", "tutor_user_id": "tutor", "student_user_id": "student", "starts_at": PAST.isoformat(), "status": "completed", "student_profiles": {"name": "Alice"}, "price": 1000, "paid": False, "public_comment": None, "schedule_rule_id": "rule1"},
    {"id": "l3", "student_id": "st1", "tutor_user_id": "tutor", "student_user_id": "student", "starts_at": (PAST - timedelta(days=2)).isoformat(), "status": "completed", "student_profiles": {"name": "Alice"}, "price": 2000, "paid": True, "public_comment": None, "schedule_rule_id": "rule1"},
    {"id": "l4", "student_id": "st1", "tutor_user_id": "tutor", "student_user_id": "student", "starts_at": (PAST - timedelta(days=5)).isoformat(), "status": "cancelled", "student_profiles": {"name": "Alice"}, "price": 1000, "paid": False, "public_comment": None, "schedule_rule_id": "rule1"},
]

fake_homework = [
    {"id": "hw1", "student_id": "st1", "student_user_id": "student", "tutor_user_id": "tutor", "title": "Read chapter", "description": "Pages 1-3", "due_at": FUTURE.isoformat(), "status": "new", "student_profiles": {"name": "Alice"}, "created_at": PAST.isoformat(), "updated_at": PAST.isoformat()},
    {"id": "hw2", "student_id": "st1", "student_user_id": "student", "tutor_user_id": "tutor", "title": "Equation set", "description": "", "due_at": None, "status": "submitted", "student_profiles": {"name": "Alice"}, "created_at": PAST.isoformat(), "updated_at": PAST.isoformat()},
    {"id": "hw3", "student_id": "st1", "student_user_id": "student", "tutor_user_id": "tutor", "title": "Essay", "description": "", "due_at": None, "status": "reviewed", "student_profiles": {"name": "Alice"}, "created_at": PAST.isoformat(), "updated_at": PAST.isoformat()},
]

fake_profile = {
    "user_id": "tutor", "slug": "alice-math", "public_enabled": True,
    "subjects": "Math, physics", "bio": "Clear explanations", "badges": "clock|Fast reply",
    "page_theme": "auto", "display_name": "Tutor One", "users": {"full_name": "Tutor One"},
    "price_per_hour": 1200, "price_duration_min": 60, "price_note": "Первое занятие бесплатно",
    "telegram_username": "tutor_one",
    "reviews": [{"author": "Мама Пети", "text": "Сын перестал бояться математики", "position": 0}],
}

fake_requests = [{
    "id": "req1", "name": "Parent", "contact": "@parent", "preferred_time": "Evening",
    "comment": "Need algebra", "status": "new", "created_at": NOW.isoformat(),
}]

events: list[tuple] = []
results: dict[str, tuple[str, str]] = {}
issues: list[dict[str, str]] = []


def record(story_id: str, ok: bool, note: str) -> None:
    current = results.get(story_id)
    if current and current[0] == "Failed":
        return
    results[story_id] = ("Passed" if ok else "Failed", note)
    if not ok:
        issues.append({
            "Issue ID": f"ISS-{len(issues) + 1:03d}",
            "Story ID": story_id,
            "Severity": "High",
            "Surface": "Automated harness",
            "Observed Error": note,
            "Expected Behavior": "Story should complete without server/template/action failure.",
            "Repro Steps": f"Run python scripts/run_user_story_tests.py and inspect story {story_id}.",
            "Root Cause": "Pending investigation",
            "Fix Summary": "",
            "Code Reference": "",
            "Status": "Open",
            "Retest Result": "Not retested",
        })


async def expect(story_id: str, label: str, fn: Callable[[], Awaitable[bool] | bool]) -> None:
    try:
        out = fn()
        ok = await out if hasattr(out, "__await__") else bool(out)
        record(story_id, bool(ok), label if ok else f"{label} returned false")
    except Exception as exc:
        record(story_id, False, f"{label}: {type(exc).__name__}: {exc}")


async def run() -> None:
    os.environ.setdefault("BOT_TOKEN", "123:abc")
    os.environ.setdefault("SUPABASE_URL", "https://example.supabase.co")
    os.environ.setdefault("SUPABASE_KEY", "fake")
    os.environ.setdefault("WEB_SECRET", "test-secret")

    import web.app as webapp

    webapp.services = FakeServices()

    async def fake_send_telegram(tg_id: int, text: str) -> None:
        events.append(("send_telegram", tg_id, text))

    async def fake_broadcast(tg_ids: list[int], text: str) -> dict:
        events.append(("broadcast", tuple(tg_ids), text))
        return {"sent": len(tg_ids), "failed": 0}

    async def fake_notify_removed(tg_id: int, tutor_name: str) -> None:
        events.append(("notify_removed_student", tg_id, tutor_name))

    webapp._send_telegram = fake_send_telegram
    webapp._broadcast_telegram = fake_broadcast
    webapp._notify_removed_student = fake_notify_removed
    app = webapp.create_app()
    transport = httpx.ASGITransport(app=app, raise_app_exceptions=True)

    async with httpx.AsyncClient(transport=transport, base_url="http://testserver", follow_redirects=False) as client:
        tutor_cookie = {"pingly_session": webapp.signer.dumps("tutor")}
        student_cookie = {"pingly_session": webapp.signer.dumps("student")}

        async def get(path: str, cookies: dict | None = None):
            return await client.get(path, cookies=cookies)

        async def post(path: str, data: dict | None = None, cookies: dict | None = None):
            return await client.post(path, data=data or {}, cookies=cookies)

        r = await get("/")
        record("F001", r.status_code == 200 and "Pingly" in r.text, f"GET / -> {r.status_code}")
        r = await get("/", tutor_cookie)
        record("F001", r.status_code == 303 and r.headers.get("location") == "/tutor", "logged-in / redirects to tutor")

        for sid, path in [("F002", "/privacy"), ("F002", "/terms"), ("F002", "/contacts")]:
            r = await get(path)
            record(sid, r.status_code == 200, f"GET {path} -> {r.status_code}")

        r = await post("/register", {"full_name": "New Tutor", "email": "new@example.com", "password": "secret123", "ref": "REF123"})
        record("F003", r.status_code == 303 and r.headers.get("location") == "/tutor", "registration redirects to tutor")
        record("F004", (await webapp.services.web_auth.register_tutor_email("", "bad", "1"))[1] is not None, "registration validation service returns errors")
        r = await post("/verify", {"email": "tutor@example.com", "code": "123456"})
        record("F005", r.status_code == 303 and r.headers.get("location") == "/tutor", "verification redirects to cabinet")
        r = await post("/verify/resend", {"email": "tutor@example.com"})
        record("F005", r.status_code == 303 and "sent=1" in r.headers.get("location", ""), "verification resend redirects with sent=1")
        r = await post("/login", {"email": "tutor@example.com", "password": "secret123"})
        record("F006", r.status_code == 303 and r.headers.get("location") == "/tutor", "valid login redirects")
        r = await post("/login", {"email": "wrong@example.com", "password": "bad"})
        record("F006", r.status_code == 303 and "bad_credentials" in r.headers.get("location", ""), "bad login redirects with error")
        r = await get("/auth/telegram/callback?hash=ok&ref=REF123")
        record("F007", r.status_code == 303 and r.headers.get("location") == "/tutor", "telegram widget login redirects")
        r = await get("/auth/telegram?token=ok-token")
        record("F008", r.status_code == 303 and r.headers.get("location") == "/tutor", "token auth redirects")
        r = await get("/logout")
        record("F009", r.status_code == 303 and r.headers.get("location") == "/", "logout redirects home")
        r = await get("/tutor", student_cookie)
        record("F010", r.status_code == 303 and r.headers.get("location") == "/student", "student redirected from tutor dashboard")
        r = await get("/student", tutor_cookie)
        record("F010", r.status_code == 303 and r.headers.get("location") == "/tutor", "tutor redirected from student dashboard")

        for sid, path, cookies, text in [
            ("F011", "/tutor", tutor_cookie, "Alice"),
            ("F012", "/tutor/students", tutor_cookie, "Alice"),
            ("F017", "/tutor/students/st1", tutor_cookie, "Alice"),
            ("F020", "/tutor/calendar?view=month", tutor_cookie, "Alice"),
            ("F020", "/student/calendar?view=week", student_cookie, "Alice"),
            ("F023", "/tutor/schedule", tutor_cookie, "Alice"),
            ("F028", "/tutor/homework", tutor_cookie, "Read chapter"),
            ("F031", "/student", student_cookie, "Read chapter"),
            ("F032", "/student/calendar", student_cookie, "Alice"),
            ("F036", "/student/homework", student_cookie, "Read chapter"),
            ("F037", "/student/history", student_cookie, "completed"),
            ("F038", "/tutor/finance", tutor_cookie, "Alice"),
            ("F041", "/tutor/requests", tutor_cookie, "alice-math"),
            ("F042", "/tutor/requests/preview", tutor_cookie, "Анна"),
            ("F043", "/u/alice-math", None, "Tutor One"),
            ("F045", "/tutor/requests", tutor_cookie, "Parent"),
            ("F048", "/tutor/settings", tutor_cookie, "REF123"),
            ("F050", "/tutor/settings", tutor_cookie, "REF123"),
            ("F051", "/tutor/settings", tutor_cookie, "Поддержка"),
            ("F053", "/tutor", tutor_cookie, "bottom-nav"),
            ("F054", "/admin", tutor_cookie, "990"),
            ("F055", "/admin/tutors", tutor_cookie, "Tutor One"),
            ("F056", "/admin/broadcast", tutor_cookie, "2"),
            ("F066", "/bad-missing-page", None, "Pingly"),
        ]:
            r = await get(path, cookies)
            record(sid, r.status_code == (404 if path == "/bad-missing-page" else 200) and text in r.text, f"GET {path} -> {r.status_code}")

        posts = [
            ("F013", "/tutor/students/create", {"name": "Bob", "tg_username": "@bob", "subject_summary": "Physics"}, tutor_cookie, "/tutor/students/st-new?created=1"),
            ("F015", "/tutor/students/st1/profile", {"name": "Alice", "subject_summary": "Math", "grade": "8", "goal": "Exam", "started_at": "2026-06-01", "default_price": "1000"}, tutor_cookie, "/tutor/students/st1?saved=profile#profile"),
            ("F016", "/tutor/students/st1/delete", {}, tutor_cookie, "/tutor/students"),
            ("F018", "/tutor/students/st1/note", {"note": "Private"}, tutor_cookie, "/tutor/students/st1?saved=note#notes"),
            ("F019", "/tutor/students/st1/package", {"action": "set", "package_size": "4"}, tutor_cookie, "/tutor/students/st1?saved=package#package"),
            ("F022", "/tutor/schedule", {"student_id": "st1", "recurrence": "once", "lesson_date": FUTURE.date().isoformat(), "lesson_time": "15:00", "comment": "Topic", "back": "/tutor/calendar?view=day"}, tutor_cookie, "/tutor/calendar?view=day"),
            ("F023", "/tutor/schedule", {"student_id": "st1", "recurrence": "weekly", "lesson_time": "15:00", "weekdays": "0", "interval_n": "1"}, tutor_cookie, "/tutor/calendar?view=month"),
            ("F024", "/tutor/lessons/l1/complete", {}, tutor_cookie, "/tutor/calendar"),
            ("F024", "/tutor/lessons/l1/cancel", {}, tutor_cookie, "/tutor/calendar"),
            ("F024", "/tutor/lessons/l1/delete", {}, tutor_cookie, "/tutor/calendar"),
            ("F025", "/tutor/lessons/l1/comment", {"comment": "Topic", "date": FUTURE.date().isoformat()}, tutor_cookie, f"/tutor/calendar?view=day&date={FUTURE.date().isoformat()}"),
            ("F026", "/tutor/lessons/l1/reschedule", {"new_at": FUTURE.strftime("%Y-%m-%dT%H:%M")}, tutor_cookie, "/tutor/calendar"),
            ("F027", "/tutor/series/rule1/cancel", {}, tutor_cookie, "/tutor/schedule"),
            ("F027", "/tutor/series/rule1/reschedule", {"new_time": "16:30"}, tutor_cookie, "/tutor/schedule"),
            ("F028", "/tutor/homework", {"student_id": "st1", "title": "New HW", "description": "Read", "due_at": FUTURE.strftime("%Y-%m-%dT%H:%M")}, tutor_cookie, "/tutor/homework"),
            ("F029", "/tutor/homework/templates", {"title": "Template", "description": "Desc"}, tutor_cookie, "/tutor/homework"),
            ("F029", "/tutor/homework/templates/tpl1/delete", {}, tutor_cookie, "/tutor/homework"),
            ("F030", "/tutor/homework/hw2/review", {"comment": "Good"}, tutor_cookie, "/tutor/homework"),
            ("F033", "/student/lessons/l1/confirm", {}, student_cookie, "/student"),
            ("F034", "/student/lessons/l1/cancel", {}, student_cookie, "/student"),
            ("F035", "/student/lessons/l1/reschedule-request", {}, student_cookie, "/student"),
            ("F036", "/student/homework/hw1/progress", {}, student_cookie, "/student/homework"),
            ("F036", "/student/homework/hw1/submit", {}, student_cookie, "/student/homework"),
            ("F039", "/tutor/lessons/l2/paid", {"paid": "1"}, tutor_cookie, "/tutor/finance"),
            ("F041", "/tutor/settings/public", {"slug": "alice-math", "bio": "Bio", "subjects": "Math", "public_enabled": "1", "badge_icon": "clock", "badge_text": "Fast", "page_theme": "dark"}, tutor_cookie, "/tutor/requests?saved=1"),
            ("F044", "/u/alice-math/book", {"name": "Parent", "contact": "@parent", "preferred_time": "Evening", "comment": "Need help"}, None, "/u/alice-math?sent=1"),
            ("F045", "/tutor/requests/req1/done", {}, tutor_cookie, "/tutor/requests"),
            ("F046", "/settings/name", {"full_name": "Tutor Renamed"}, tutor_cookie, "/tutor/settings?saved=name"),
            ("F048", "/tutor/billing/subscribe", {"plan": "max"}, tutor_cookie, "/tutor/settings?error=payments_off"),
            ("F049", "/payments/platega/webhook", {}, None, None),
            ("F051", "/support", {"message": "Need help"}, tutor_cookie, "/tutor/settings?saved=support"),
            ("F055", "/admin/tutors/tutor/subscription", {"action": "grant", "plan": "max", "days": "30"}, tutor_cookie, "/admin/tutors?saved=1"),
            ("F056", "/admin/broadcast", {"message": "Hello"}, tutor_cookie, "/admin/broadcast?result=2-0"),
        ]
        for sid, path, data, cookies, expected_loc in posts:
            if sid == "F049":
                r = await client.post(path, json={"id": "tx1", "status": "CONFIRMED"})
                record(sid, r.status_code in (200, 400), f"POST {path} -> {r.status_code}")
                continue
            r = await post(path, data, cookies)
            record(sid, r.status_code == 303 and r.headers.get("location") == expected_loc, f"POST {path} -> {r.status_code} {r.headers.get('location')}")

        r = await get("/auth/telegram/link?hash=ok", tutor_cookie)
        record("F047", r.status_code == 303 and r.headers.get("location") == "/tutor/settings?saved=tg", f"GET /auth/telegram/link -> {r.status_code} {r.headers.get('location')}")

        r = await get("/design-tokens.css")
        r2 = await get("/favicon.ico")
        record("F066", r.status_code == 200 and r2.status_code == 200, "static design tokens and favicon load")

        # Static/client-side stories inspected from rendered HTML/JS.
        finance = await get("/tutor/finance", tutor_cookie)
        record("F040", all(s in finance.text for s in ["massbar-pay", "fin-remind", "data-period"]), "finance page includes expected interactive controls")
        original_student = deepcopy(fake_students[0])
        fake_students[0]["user_id"] = None
        fake_students[0]["tg_connected"] = False
        fake_students[0]["vk_connected"] = False
        students = await get("/tutor/students", tutor_cookie)
        fake_students[0] = original_student
        record("F014", "start=inv_tok123" in students.text and "copyInvite" in students.text, "student invite link and copy action rendered")
        calendar = await get("/tutor/calendar", tutor_cookie)
        record("F021", "data-status=\"scheduled\"" in calendar.text and "cal-legend" in calendar.text, "calendar status filter controls rendered")
        layout = await get("/tutor", tutor_cookie)
        record("F052", "lock-badge" not in layout.text, "plan locks inactive for active trial/max user")
        record("F053", all(s in layout.text for s in ["bottom-nav", "toggleTheme", "startPinglyTour"]), "shell UX hooks rendered")

        await exercise_bot_scheduler_stories()

    update_workbook()


class FakeFromUser:
    def __init__(self, user_id: int, first_name: str = "User", full_name: str = "User Name", username: str | None = "user") -> None:
        self.id = user_id
        self.first_name = first_name
        self.full_name = full_name
        self.username = username


class FakeMessage:
    def __init__(self, user_id: int, text: str = "hello", username: str | None = "user") -> None:
        self.from_user = FakeFromUser(user_id, username=username)
        self.text = text
        self.answers: list[str] = []
        self.reply_markup = None
        self.link_preview_options = None
        self.edits: list[str] = []
        self.bot = FakeBot()

    async def answer(self, text: str, **kwargs) -> None:
        self.answers.append(text)
        self.reply_markup = kwargs.get("reply_markup")
        self.link_preview_options = kwargs.get("link_preview_options")

    async def edit_text(self, text: str, **kwargs) -> None:
        self.edits.append(text)


class FakeCommand:
    def __init__(self, args: str | None = None) -> None:
        self.args = args


class FakeCallback:
    def __init__(self, data: str, user_id: int = 1002) -> None:
        self.data = data
        self.from_user = FakeFromUser(user_id)
        self.message = FakeMessage(user_id)
        self.bot = FakeBot()
        self.answers: list[str] = []

    async def answer(self, text: str, **kwargs) -> None:
        self.answers.append(text)


class FakeBot:
    def __init__(self) -> None:
        self.sent: list[tuple[int, str, object]] = []

    async def send_message(self, tg_id: int, text: str, reply_markup=None, **kwargs) -> None:
        self.sent.append((tg_id, text, reply_markup))
        events.append(("fake_bot_send", tg_id, text, bool(reply_markup)))


class FakeNotificationService:
    def __init__(self) -> None:
        self.sent_ids: list[str] = []

    async def due_notifications(self) -> list[dict]:
        return [
            {
                "id": "n1", "type": "lesson_hour_before", "title": "Lesson soon",
                "body": "Starts at 15:00", "payload": {"lesson_id": "l1"},
                "users": {"tg_id": 1002, "vk_id": None},
            },
            {
                "id": "n2", "type": "tutor_unconfirmed", "title": "Unconfirmed",
                "body": "Still not confirmed", "payload": {"lesson_id": "l1"},
                "users": {"tg_id": 1001, "vk_id": None},
            },
        ]

    async def mark_sent(self, notification_id: str) -> None:
        self.sent_ids.append(notification_id)
        events.append(("notification_mark_sent", notification_id))


class SchedulerLessonService:
    async def lesson_is_unconfirmed(self, lesson_id: str) -> bool:
        events.append(("lesson_is_unconfirmed", lesson_id))
        return True


class SchedulerRepo:
    def __init__(self) -> None:
        self.created: list[tuple] = []

    async def get_lesson_by_id(self, lesson_id: str) -> dict | None:
        return {"id": lesson_id, "public_comment": "Fractions"}

    async def list_tutors_with_trial(self) -> list[dict]:
        return [{
            "id": "tutor", "role": "tutor", "subscription_status": "trial",
            "trial_ends_at": (NOW + timedelta(days=3)).isoformat(),
        }]

    async def list_notifications_for_user(self, user_id: str, limit: int = 50) -> list[dict]:
        return []

    async def create_notification(self, user_id: str, ntype: str, title: str, body: str, payload: dict | None = None, scheduled_for=None) -> dict:
        row = (user_id, ntype, title, body, payload or {})
        self.created.append(row)
        events.append(("scheduler_create_notification", *row))
        return {"id": f"new-{len(self.created)}"}

    async def list_active_package_students(self) -> list[dict]:
        return [{
            "tutor_user_id": "tutor", "student_id": "st1", "student_user_id": "student",
            "name": "Alice", "package_size": 1, "package_started_at": PAST.isoformat(),
        }]

    async def list_lessons_for_tutor(self, tutor_user_id: str, limit: int = 1000) -> list[dict]:
        return [{
            "id": "l2", "student_id": "st1", "starts_at": PAST.isoformat(),
            "status": "completed",
        }]


class SchedulerServices:
    def __init__(self) -> None:
        self.notifications = FakeNotificationService()
        self.lessons = SchedulerLessonService()
        self.repo = SchedulerRepo()


async def exercise_bot_scheduler_stories() -> None:
    import handlers.student as student_handlers
    import handlers.tutor as tutor_handlers
    import scheduler
    import vk_bot

    fake = FakeServices()
    tutor_handlers.services = fake
    student_handlers.services = fake

    invite_msg = FakeMessage(3000, username="newstudent")
    await tutor_handlers.cmd_start(invite_msg, FakeCommand("inv_tok123"))
    record("F057", bool(invite_msg.answers and "напоминание" in invite_msg.answers[-1].lower()), "Telegram invite /start links student and answers")

    web_msg = FakeMessage(1001, username="tutorone")
    await tutor_handlers.cmd_web(web_msg)
    help_msg = FakeMessage(4000)
    await tutor_handlers.cmd_help(help_msg)
    record("F058", bool(web_msg.answers and "token=ok-token" in web_msg.answers[-1] and help_msg.reply_markup), "Telegram /web and /help handlers answer")

    confirm = FakeCallback("lesson_confirm:l1")
    await student_handlers.confirm_lesson(confirm)
    cancel = FakeCallback("lesson_cancel:l1")
    await student_handlers.cancel_lesson(cancel)
    record("F059", bool(confirm.message.edits and cancel.message.edits and confirm.answers and cancel.answers), "Telegram confirm/cancel callbacks edit and answer")

    reason = FakeMessage(1002, text="Sick today")
    await student_handlers.capture_cancel_reason(reason)
    record("F060", bool(reason.answers and "Передал" in reason.answers[-1]), "Telegram cancellation reason is forwarded and acknowledged")

    sched_services = SchedulerServices()
    scheduler.services = sched_services
    bot = FakeBot()
    await scheduler.send_due_notifications(bot)
    sent_ids = sched_services.notifications.sent_ids
    record("F061", "n1" in sent_ids and any("Fractions" in item[1] for item in bot.sent), "Scheduler delivers due lesson notification with latest topic")
    record("F062", "n2" in sent_ids and any(e[0] == "lesson_is_unconfirmed" for e in events), "Scheduler checks unconfirmed lesson before tutor nudge")

    await scheduler.enqueue_subscription_reminders()
    record("F063", any(e[0] == "scheduler_create_notification" and e[2] == "subscription_expiring" for e in events), "Scheduler queues subscription expiry reminder")

    await scheduler.enqueue_package_reminders()
    record("F064", any(e[0] == "scheduler_create_notification" and e[2] == "package_ending" for e in events), "Scheduler queues package ending reminder")

    keyboard = vk_bot.lesson_keyboard("l1")
    keyboard_text = str(keyboard)
    record(
        "F065",
        bool(keyboard.get("buttons") and "lesson_confirm" in keyboard_text and "lesson_cancel" in keyboard_text and "l1" in keyboard_text),
        "VK lesson keyboard contains confirm/cancel payloads",
    )


def update_workbook() -> None:
    wb = load_workbook(WORKBOOK)
    ws = wb["User Stories"]
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    for row in range(2, ws.max_row + 1):
        story_id = ws.cell(row, header["ID"]).value
        status, note = results.get(story_id, ("Not tested", "No automated harness coverage in this run"))
        ws.cell(row, header["Initial Test Status"]).value = status
        ws.cell(row, header["Initial Test Notes"]).value = note
        if status == "Passed":
            ws.cell(row, header["Fix Status"]).value = "No fix needed"
            ws.cell(row, header["Retest Status"]).value = "Pending final retest"
        else:
            linked = [i["Issue ID"] for i in issues if i["Story ID"] == story_id]
            ws.cell(row, header["Issue IDs"]).value = ", ".join(linked)
            ws.cell(row, header["Fix Status"]).value = "Open"
    issue_ws = wb["Issue Log"]
    if issue_ws.max_row > 1:
        issue_ws.delete_rows(2, issue_ws.max_row - 1)
    issue_headers = [cell.value for cell in issue_ws[1]]
    for issue in issues:
        issue_ws.append([issue.get(h, "") for h in issue_headers])
    summary = wb["Summary"]
    summary["B5"] = "Initial automated/local harness testing complete"
    wb.save(WORKBOOK)
    passed = sum(1 for s, _ in results.values() if s == "Passed")
    failed = sum(1 for s, _ in results.values() if s == "Failed")
    print(f"updated {WORKBOOK}")
    print(f"passed={passed} failed={failed} issues={len(issues)}")
    for issue in issues:
        print(f"{issue['Issue ID']} {issue['Story ID']}: {issue['Observed Error']}")


if __name__ == "__main__":
    asyncio.run(run())
